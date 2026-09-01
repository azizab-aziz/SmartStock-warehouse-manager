"""
export_report.py - reads a CSV written by the C app (already sorted by
category, then product name) and produces a styled .xlsx report grouped
into one section per category, plus a top-level summary band.
No pandas/numpy - just csv (stdlib) + openpyxl.

Usage:
    python export_report.py [csv_path] [xlsx_path] [sheet_name]
"""

import csv
import os
import subprocess
import sys

DEFAULT_CSV = "exports/export.csv"
DEFAULT_XLSX = "exports/rapport_stock.xlsx"
DEFAULT_SHEET = "Rapport Stock"

HEADER_COLOR = "1E293B"     # navy - matches SmartStock's header bar
SUBHEADER_COLOR = "334155"  # lighter navy - column header row inside a section
ROW_ALT_COLOR = "EFF6FF"    # very light blue - matches the app's row hover tint
SUMMARY_COLOR = "0F172A"    # darkest navy - top summary band
GREEN = "10B981"
RED = "DC2626"

# Columns written by inv_export_csv() (in this exact order). "Categorie"
# is used to split into sections, not shown as its own column.
CSV_COLUMNS = ["PRD", "SKU", "Nom", "Categorie", "Unite",
               "Quantite", "Prix_Unitaire", "Valeur_Stock", "Seuil_Alerte", "Statut"]
DISPLAY_COLUMNS = ["PRD", "SKU", "Nom", "Unite", "Quantite", "Prix_Unitaire", "Valeur_Stock", "Seuil_Alerte", "Statut"]
NUMERIC_COLUMNS = {"Quantite", "Prix_Unitaire", "Valeur_Stock", "Seuil_Alerte"}


def ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("openpyxl introuvable, installation en cours...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])


ensure_openpyxl()

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def to_number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        return [], []
    return rows[0], rows[1:]


def group_by_category(headers, data):
    """Splits rows into (category_name, [rows]) groups, preserving the
    order they arrive in - which is already category-then-name sorted
    on the C side, so no re-sorting needed here."""
    cat_idx = headers.index("Categorie")
    groups = []
    current_cat, current_rows = None, []
    for row in data:
        cat = row[cat_idx] if cat_idx < len(row) else ""
        if cat != current_cat and current_rows:
            groups.append((current_cat, current_rows))
            current_rows = []
        current_cat = cat
        current_rows.append(row)
    if current_rows:
        groups.append((current_cat, current_rows))
    return groups


def build_workbook(headers, data, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or DEFAULT_SHEET)[:31]

    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor=ROW_ALT_COLOR)
    col_count = len(DISPLAY_COLUMNS)

    qty_idx = CSV_COLUMNS.index("Quantite")
    price_idx = CSV_COLUMNS.index("Prix_Unitaire")
    value_idx = CSV_COLUMNS.index("Valeur_Stock")
    threshold_idx = CSV_COLUMNS.index("Seuil_Alerte")
    statut_idx = CSV_COLUMNS.index("Statut")

    # ---- top summary band ----
    total_products = len(data)
    total_qty = sum((to_number(r[qty_idx]) or 0) for r in data if qty_idx < len(r))
    total_value = sum((to_number(r[value_idx]) or 0) for r in data if value_idx < len(r))
    groups = group_by_category(headers, data)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value="Resume general")
    title_cell.fill = PatternFill("solid", fgColor=SUMMARY_COLOR)
    title_cell.font = Font(color="FFFFFF", bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    summary_labels = ["Categories", "Produits", "Unites en stock", "Valeur totale du stock"]
    summary_values = [len(groups), total_products, int(total_qty), f"{total_value:.2f} DT"]
    for ci, (label, value) in enumerate(zip(summary_labels, summary_values), start=1):
        lc = ws.cell(row=2, column=ci, value=label)
        lc.font = Font(bold=True, color=HEADER_COLOR)
        lc.border = border
        vc = ws.cell(row=3, column=ci, value=value)
        vc.font = Font(size=12, bold=True, color=GREEN)
        vc.border = border
    row_cursor = 5

    # ---- one section per category ----
    for cat_name, rows in groups:
        display_name = cat_name if cat_name else "Sans categorie"

        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=col_count)
        band = ws.cell(row=row_cursor, column=1, value=display_name)
        band.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        band.font = Font(color="FFFFFF", bold=True, size=12)
        band.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_cursor].height = 22
        row_cursor += 1

        for ci, h in enumerate(DISPLAY_COLUMNS, start=1):
            c = ws.cell(row=row_cursor, column=ci, value=h)
            c.fill = PatternFill("solid", fgColor=SUBHEADER_COLOR)
            c.font = Font(color="FFFFFF", bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        row_cursor += 1

        cat_qty_total = 0.0
        cat_value_total = 0.0

        for i, row in enumerate(rows):
            is_alt = i % 2 == 1
            low_stock = False
            q = to_number(row[qty_idx]) if qty_idx < len(row) else None
            s = to_number(row[threshold_idx]) if threshold_idx < len(row) else None
            if q is not None and s is not None and q <= s:
                low_stock = True
            if q is not None:
                cat_qty_total += q
            v = to_number(row[value_idx]) if value_idx < len(row) else None
            if v is not None:
                cat_value_total += v

            for ci, h in enumerate(DISPLAY_COLUMNS, start=1):
                src_idx = CSV_COLUMNS.index(h)
                raw = row[src_idx] if src_idx < len(row) else ""
                value = raw
                if h in NUMERIC_COLUMNS:
                    num = to_number(raw)
                    value = num if num is not None else raw  # fallback: keep as text, don't crash

                cell = ws.cell(row=row_cursor, column=ci, value=value)
                cell.border = border
                if is_alt:
                    cell.fill = alt_fill

                if h == "Quantite" and low_stock:
                    cell.font = Font(color=RED, bold=True)
                elif h == "Valeur_Stock":
                    cell.font = Font(color=GREEN)
                elif h == "Statut":
                    cell.font = Font(color=RED if low_stock else GREEN, bold=True)

            row_cursor += 1

        # subtotal line for this category
        sub_label = ws.cell(row=row_cursor, column=1, value=f"Sous-total ({len(rows)} produit(s))")
        sub_label.font = Font(bold=True, color=HEADER_COLOR)
        qty_col = DISPLAY_COLUMNS.index("Quantite") + 1
        value_col = DISPLAY_COLUMNS.index("Valeur_Stock") + 1
        qc = ws.cell(row=row_cursor, column=qty_col, value=int(cat_qty_total))
        qc.font = Font(bold=True)
        vc = ws.cell(row=row_cursor, column=value_col, value=round(cat_value_total, 2))
        vc.font = Font(bold=True, color=GREEN)
        for ci in range(1, col_count + 1):
            ws.cell(row=row_cursor, column=ci).border = border
        row_cursor += 2  # blank spacer row before next category

    # ---- column widths: longest content across the whole sheet + margin ----
    for ci, h in enumerate(DISPLAY_COLUMNS, start=1):
        max_len = len(h)
        src_idx = CSV_COLUMNS.index(h)
        for row in data:
            if src_idx < len(row):
                max_len = max(max_len, len(str(row[src_idx])))
        ws.column_dimensions[get_column_letter(ci)].width = max_len + 4

    return wb


def main():
    csv_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_CSV
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_XLSX
    sheet_name = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_SHEET

    if not os.path.isfile(csv_path):
        print(f"Erreur: fichier CSV introuvable: {csv_path}")
        sys.exit(1)

    headers, data = read_csv(csv_path)
    if not headers:
        print("Erreur: CSV vide")
        sys.exit(1)

    wb = build_workbook(headers, data, sheet_name)
    out_dir = os.path.dirname(xlsx_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(xlsx_path)
    print("Excel exported successfully")


if __name__ == "__main__":
    main()